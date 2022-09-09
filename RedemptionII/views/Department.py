from openpyxl import load_workbook

from django.shortcuts import render, redirect, HttpResponse

from RedemptionII import models
from RedemptionII.utils.pagination import Pagination
from RedemptionII.utils.form import UserModelForm, PrettyModelForm, PrettyEditModelForm


# Create your views here.


def depart_list(request):
    """部门列表"""

    # 获取数据库中所有的部门列表
    # [对象， 对象， 对象]
    queryset = models.Department.objects.all()
    page_object = Pagination(request, queryset, page_size=10)
    context = {
        "queryset": page_object.page_queryset,
        "page_string": page_object.html()
    }

    return render(request, 'depart_list.html', context)


def depart_add(request):
    """添加部门"""

    if request.method == "GET":
        return render(request, 'depart_add.html')

    # 获取用户 POST 提交过来的数据
    title = request.POST.get("title")

    # 保存到数据库
    models.Department.objects.create(title=title)

    # 重新定向回部门列表
    return redirect("/depart/list/")


def depart_delete(request):
    """删除部门"""

    # 获取ID
    nid = request.GET.get('nid')

    # 删除
    models.Department.objects.filter(id=nid).delete()

    # 重新定向回部门列表
    return redirect("/depart/list")


def depart_edit(request, nid):
    """修改部门"""

    if request.method == "GET":
        # 根据nid，获取它的数据 [obj, obj...]
        row_object = models.Department.objects.filter(id=nid).first()

        return render(request, "depart_edit.html", {"row_object": row_object})

    # 获取用户更改的标题
    title = request.POST.get("title")

    # 根据 ID 找到数据库中的数据并进行更新
    models.Department.objects.filter(id=nid).update(title=title)

    # 重新定向回部门列表
    return redirect("/depart/list/")


def depart_multi(request):
    """批量上传 excel 文件"""

    # 获取用户上传的文件
    file_object = request.FILES.get('exc')
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]

    # 循环获取每一行数据
    for row in sheet.iter_rows(min_row=2):
        text = row[0].value
        exists = models.Department.objects.filter(title=text).exists()
        if not exists:
            models.Department.objects.create(title=text)

    return redirect('/depart/list/')
